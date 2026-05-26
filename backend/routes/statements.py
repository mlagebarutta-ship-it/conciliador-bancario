from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse
from typing import Dict, List, Optional
from datetime import datetime, timezone
import uuid
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from database import db
from models.transaction import (
    BankStatement, Transaction, TransactionUpdate, BulkUpdateRequest, ClassificationHistory
)
from auth.helpers import require_auth, get_tenant_id, get_user_allowed_company_ids, log_activity
from services.parsers import parse_ofx_statement, parse_pdf_statement, parse_excel_statement
from services.classification import classify_transaction
from utils.helpers import clean_cnpj

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")


@router.post("/bank-statements/upload")
async def upload_statement(
    file: UploadFile = File(...),
    company_id: str = Query(...),
    chart_id: str = Query(...),
    bank_name: str = Query(...),
    period: str = Query(...),
    current_user: Dict = Depends(require_auth)
):
    tenant_id = get_tenant_id(current_user)
    allowed_company_ids = await get_user_allowed_company_ids(current_user)
    if allowed_company_ids is not None and company_id not in allowed_company_ids:
        raise HTTPException(status_code=403, detail="Você não tem acesso a esta empresa")
    company_query = {"id": company_id}
    if tenant_id:
        company_query["tenant_id"] = tenant_id
    company = await db.companies.find_one(company_query, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    content = await file.read()
    transactions = []
    if file.filename.endswith('.ofx'):
        transactions = parse_ofx_statement(content)
    elif file.filename.endswith('.pdf'):
        transactions = parse_pdf_statement(content)
    elif file.filename.endswith(('.xlsx', '.xls', '.csv')):
        transactions = parse_excel_statement(content)
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Use OFX, PDF, Excel ou CSV")

    statement_id = str(uuid.uuid4())
    classified_transactions = []
    for trans in transactions:
        classification = await classify_transaction(
            trans['description'], trans['amount'], trans['transaction_type'], chart_id, company_id
        )
        trans_obj = Transaction(
            id=str(uuid.uuid4()), statement_id=statement_id,
            date=trans['date'], description=trans['description'],
            document=trans.get('document'), amount=abs(trans['amount']),
            transaction_type=trans['transaction_type'],
            debit_account=classification['debit_account'],
            credit_account=classification['credit_account'],
            status=classification['status']
        )
        classified_transactions.append(trans_obj)

    total_inflows = sum(t.amount for t in classified_transactions if t.transaction_type == 'C')
    total_outflows = sum(t.amount for t in classified_transactions if t.transaction_type == 'D')
    classified_count = sum(1 for t in classified_transactions if t.status == 'CLASSIFICADO')
    manual_count = len(classified_transactions) - classified_count

    statement = BankStatement(
        id=statement_id, tenant_id=tenant_id, company_id=company_id, chart_id=chart_id,
        filename=file.filename, bank_name=bank_name, period=period,
        total_transactions=len(classified_transactions),
        classified_count=classified_count, manual_count=manual_count,
        total_inflows=total_inflows, total_outflows=total_outflows,
        balance=total_inflows - total_outflows, status='COMPLETED',
        processed_at=datetime.now(timezone.utc)
    )
    statement_doc = statement.model_dump()
    statement_doc['created_at'] = statement_doc['created_at'].isoformat()
    statement_doc['processed_at'] = statement_doc['processed_at'].isoformat()
    await db.bank_statements.insert_one(statement_doc)

    for trans in classified_transactions:
        await db.transactions.insert_one(trans.model_dump())

    await log_activity(
        usuario_id=current_user['id'], usuario_nome=current_user['nome'],
        acao="Importou extrato", detalhes=f"Período: {period}, Arquivo: {file.filename}",
        empresa_id=company_id, empresa_nome=company.get('name'),
        tenant_id=tenant_id, tenant_nome=current_user.get('tenant', {}).get('nome')
    )
    return {"statement": statement, "transactions": classified_transactions}


@router.get("/bank-statements", response_model=List[BankStatement])
async def get_statements(company_id: Optional[str] = None, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    query = {}
    if tenant_id:
        query["tenant_id"] = tenant_id
    allowed_company_ids = await get_user_allowed_company_ids(current_user)
    if allowed_company_ids is not None:
        if not allowed_company_ids:
            return []
        if company_id:
            if company_id not in allowed_company_ids:
                raise HTTPException(status_code=403, detail="Você não tem acesso a esta empresa")
            query["company_id"] = company_id
        else:
            query["company_id"] = {"$in": allowed_company_ids}
    elif company_id:
        query["company_id"] = company_id
    statements = await db.bank_statements.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for s in statements:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
        if isinstance(s.get('processed_at'), str):
            s['processed_at'] = datetime.fromisoformat(s['processed_at'])
    return statements


@router.get("/bank-statements/{statement_id}", response_model=BankStatement)
async def get_statement(statement_id: str, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    query = {"id": statement_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    statement = await db.bank_statements.find_one(query, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    allowed_company_ids = await get_user_allowed_company_ids(current_user)
    if allowed_company_ids is not None and statement.get('company_id') not in allowed_company_ids:
        raise HTTPException(status_code=403, detail="Você não tem acesso a este extrato")
    if isinstance(statement.get('created_at'), str):
        statement['created_at'] = datetime.fromisoformat(statement['created_at'])
    if isinstance(statement.get('processed_at'), str):
        statement['processed_at'] = datetime.fromisoformat(statement['processed_at'])
    return statement


@router.get("/bank-statements/{statement_id}/transactions", response_model=List[Transaction])
async def get_transactions(statement_id: str, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    statement_query = {"id": statement_id}
    if tenant_id:
        statement_query["tenant_id"] = tenant_id
    statement = await db.bank_statements.find_one(statement_query, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    allowed_company_ids = await get_user_allowed_company_ids(current_user)
    if allowed_company_ids is not None and statement.get('company_id') not in allowed_company_ids:
        raise HTTPException(status_code=403, detail="Você não tem acesso a este extrato")
    transactions = await db.transactions.find({"statement_id": statement_id}, {"_id": 0}).to_list(1000)
    return transactions


@router.put("/transactions/bulk-update")
async def bulk_update_transactions(request: BulkUpdateRequest, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    if not request.transaction_ids:
        raise HTTPException(status_code=400, detail="Nenhuma transação selecionada")
    if not request.update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    valid_fields = ['debit_account', 'credit_account', 'status']
    update_data = {k: v for k, v in request.update_data.items() if k in valid_fields and v}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo válido para atualizar")

    result = await db.transactions.update_many(
        {"id": {"$in": request.transaction_ids}}, {"$set": update_data}
    )

    if update_data.get('debit_account') and update_data.get('credit_account'):
        updated_transactions = await db.transactions.find(
            {"id": {"$in": request.transaction_ids}}, {"_id": 0}
        ).to_list(len(request.transaction_ids))
        statement_ids = set(t['statement_id'] for t in updated_transactions)
        statements = {}
        for st_id in statement_ids:
            st = await db.bank_statements.find_one({"id": st_id}, {"_id": 0})
            if st:
                statements[st_id] = st
        descriptions_saved = set()
        for trans in updated_transactions:
            statement = statements.get(trans['statement_id'])
            if not statement:
                continue
            company_id = statement['company_id']
            description = trans['description']
            key = f"{company_id}:{description}"
            if key in descriptions_saved:
                continue
            descriptions_saved.add(key)
            existing = await db.classification_history.find_one({
                "company_id": company_id, "description_pattern": description
            }, {"_id": 0})
            if existing:
                await db.classification_history.update_one(
                    {"id": existing['id']},
                    {"$set": {"debit_account": update_data['debit_account'], "credit_account": update_data['credit_account'], "last_used": datetime.now(timezone.utc).isoformat()},
                     "$inc": {"usage_count": 1}}
                )
            else:
                history_entry = ClassificationHistory(
                    tenant_id=tenant_id, company_id=company_id,
                    description_pattern=description,
                    transaction_type=trans.get('transaction_type', 'D'),
                    debit_account=update_data['debit_account'],
                    credit_account=update_data['credit_account']
                )
                history_doc = history_entry.model_dump()
                history_doc['created_at'] = history_doc['created_at'].isoformat()
                history_doc['last_used'] = history_doc['last_used'].isoformat()
                await db.classification_history.insert_one(history_doc)
        logger.info(f"Histórico de classificação atualizado para {len(descriptions_saved)} descrições únicas")

    return {"message": f"{result.modified_count} transações atualizadas com sucesso", "updated_count": result.modified_count}


@router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(transaction_id: str, update: TransactionUpdate, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    trans = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if trans and update.debit_account and update.credit_account:
        statement = await db.bank_statements.find_one({"id": trans['statement_id']}, {"_id": 0})
        if statement:
            company_id = statement['company_id']
            description = trans['description']
            trans_type = trans['transaction_type']
            existing_history = await db.classification_history.find_one({
                "company_id": company_id, "description_pattern": description
            }, {"_id": 0})
            if existing_history:
                await db.classification_history.update_one(
                    {"id": existing_history['id']},
                    {"$set": {"debit_account": update.debit_account, "credit_account": update.credit_account, "last_used": datetime.now(timezone.utc).isoformat()},
                     "$inc": {"usage_count": 1}}
                )
            else:
                history_entry = ClassificationHistory(
                    tenant_id=tenant_id, company_id=company_id,
                    description_pattern=description, transaction_type=trans_type,
                    debit_account=update.debit_account, credit_account=update.credit_account
                )
                history_doc = history_entry.model_dump()
                history_doc['created_at'] = history_doc['created_at'].isoformat()
                history_doc['last_used'] = history_doc['last_used'].isoformat()
                await db.classification_history.insert_one(history_doc)
    return trans


@router.get("/bank-statements/{statement_id}/export")
async def export_statement(statement_id: str, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    query = {"id": statement_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    statement = await db.bank_statements.find_one(query, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    transactions = await db.transactions.find({"statement_id": statement_id}, {"_id": 0}).to_list(1000)
    company = await db.companies.find_one({"id": statement['company_id']}, {"_id": 0})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    headers = ['Descrição', 'Data', 'Valor', 'Conta Débito', 'Conta Crédito', 'Histórico']
    ws.append(headers)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for trans in transactions:
        valor = trans['amount'] if trans['transaction_type'] == 'C' else -trans['amount']
        ws.append([
            'IMPLANTAÇÃO DE SALDO', trans['date'], valor,
            trans.get('debit_account', ''), trans.get('credit_account', ''),
            trans['description']
        ])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    cnpj_clean = clean_cnpj(company['cnpj'])
    filename = f"{cnpj_clean}_{statement['bank_name'].upper()}_{statement['period'].replace('/', '')}_{statement_id[:8]}_LANCAMENTOS.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@router.delete("/bank-statements/{statement_id}")
async def delete_statement(statement_id: str, current_user: Dict = Depends(require_auth)):
    tenant_id = get_tenant_id(current_user)
    query = {"id": statement_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    statement = await db.bank_statements.find_one(query, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    result = await db.bank_statements.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    await db.transactions.delete_many({"statement_id": statement_id})
    await log_activity(
        usuario_id=current_user['id'], usuario_nome=current_user['nome'],
        acao="Excluiu extrato", detalhes=f"Período: {statement.get('period')}",
        empresa_id=statement.get('company_id'), tenant_id=tenant_id,
        tenant_nome=current_user.get('tenant', {}).get('nome')
    )
    return {"message": "Extrato excluído com sucesso"}
