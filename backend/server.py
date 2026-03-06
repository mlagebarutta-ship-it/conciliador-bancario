from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import ofxparse
import pdfplumber
import io
import re
from collections import defaultdict

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ============= MODELS =============

# Status de processamento contábil
PROCESSING_STATUS = {
    'NAO_INICIADO': 'Não iniciado',
    'EM_PROCESSAMENTO': 'Em processamento',
    'AGUARDANDO_DOCUMENTOS': 'Aguardando documentos',
    'CONCLUIDO': 'Concluído',
    'ARQUIVADO': 'Arquivado'
}

class AccountingProcess(BaseModel):
    """Processamento contábil mensal de uma empresa"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    company_name: str
    year: int
    month: int
    period: str  # MM/YYYY
    status: str = 'NAO_INICIADO'
    responsible: Optional[str] = None
    observations: Optional[str] = None
    statement_ids: List[str] = Field(default_factory=list)
    total_transactions: int = 0
    classified_transactions: int = 0
    pending_transactions: int = 0
    is_archived: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None

class AccountingProcessCreate(BaseModel):
    company_id: str
    year: int
    month: int
    responsible: Optional[str] = None
    observations: Optional[str] = None

class AccountingProcessUpdate(BaseModel):
    status: Optional[str] = None
    responsible: Optional[str] = None
    observations: Optional[str] = None

class Company(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cnpj: str
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanyCreate(BaseModel):
    cnpj: str
    name: str
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class ChartOfAccounts(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ChartOfAccountsCreate(BaseModel):
    company_id: str
    name: str
    description: Optional[str] = None

class AccountItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    chart_id: str
    code: str
    description: str
    account_type: str  # ATIVO, PASSIVO, RECEITA, DESPESA
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AccountItemCreate(BaseModel):
    chart_id: str
    code: str
    description: str
    account_type: str

class ClassificationRule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    keyword: str
    debit_account_code: Optional[str] = None
    credit_account_code: Optional[str] = None
    description: Optional[str] = None
    priority: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClassificationRuleCreate(BaseModel):
    keyword: str
    debit_account_code: Optional[str] = None
    credit_account_code: Optional[str] = None
    description: Optional[str] = None
    priority: int = 0

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    statement_id: str
    date: str
    description: str
    document: Optional[str] = None
    amount: float
    transaction_type: str  # C or D
    debit_account: Optional[str] = None
    credit_account: Optional[str] = None
    status: str  # CLASSIFICADO or CLASSIFICAR MANUALMENTE
    notes: Optional[str] = None

class TransactionCreate(BaseModel):
    date: str
    description: str
    document: Optional[str] = None
    amount: float
    transaction_type: str

class TransactionUpdate(BaseModel):
    date: Optional[str] = None
    description: Optional[str] = None
    document: Optional[str] = None
    amount: Optional[float] = None
    debit_account: Optional[str] = None
    credit_account: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class ClassificationHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    description_pattern: str  # Descrição original da transação
    transaction_type: str  # C ou D
    debit_account: str
    credit_account: str
    usage_count: int = 1  # Quantas vezes foi usada
    last_used: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BankStatement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    chart_id: str
    filename: str
    bank_name: str
    period: str  # MM/YYYY
    total_transactions: int
    classified_count: int
    manual_count: int
    total_inflows: float
    total_outflows: float
    balance: float
    status: str  # PROCESSING, COMPLETED, ERROR
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_at: Optional[datetime] = None

class BankStatementCreate(BaseModel):
    company_id: str
    chart_id: str
    bank_name: str
    period: str

class ProcessingResult(BaseModel):
    statement: BankStatement
    transactions: List[Transaction]

# ============= HELPER FUNCTIONS =============

def clean_cnpj(cnpj: str) -> str:
    """Remove caracteres especiais do CNPJ"""
    return re.sub(r'[^0-9]', '', cnpj)

def parse_brazilian_number(value_str: str) -> float:
    """Converte número no formato brasileiro para float"""
    if not value_str:
        return 0.0
    
    value_str = str(value_str).strip()
    
    # Ignorar cabeçalhos e textos
    if any(x in value_str.lower() for x in ['entradas', 'saídas', 'saidas', 'créditos', 'creditos', 'débitos', 'debitos', 'saldo', 'total']):
        return 0.0
    
    # Verificar se é negativo (entre parênteses, com sinal no início, ou com sinal NO FINAL)
    is_negative = False
    if value_str.startswith('(') and value_str.endswith(')'):
        is_negative = True
        value_str = value_str[1:-1]
    elif value_str.startswith('-'):
        is_negative = True
        value_str = value_str[1:]
    elif value_str.endswith('-'):  # FORMATO SANTANDER: 415,84-
        is_negative = True
        value_str = value_str[:-1]
    elif value_str.endswith('+'):
        value_str = value_str[:-1]
    
    # Remover R$ e espaços
    value_str = value_str.replace('R$', '').replace(' ', '').strip()
    
    # Detectar formato (brasileiro: 1.234,56 ou americano: 1,234.56)
    if ',' in value_str and '.' in value_str:
        # Verificar qual vem por último
        if value_str.rfind(',') > value_str.rfind('.'):
            # Formato brasileiro: 1.234,56
            value_str = value_str.replace('.', '').replace(',', '.')
        else:
            # Formato americano: 1,234.56
            value_str = value_str.replace(',', '')
    elif ',' in value_str:
        # Só tem vírgula - provavelmente brasileiro
        value_str = value_str.replace(',', '.')
    # Se só tem ponto, manter como está (formato americano)
    
    try:
        result = float(value_str)
        return -result if is_negative else result
    except:
        return 0.0

def parse_pdf_statement(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse PDF bank statement - Suporta múltiplos formatos de extrato"""
    try:
        transactions = []
        
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            # Primeiro, tentar extrair tabelas estruturadas
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Detectar colunas pelo cabeçalho
                    header = [str(h).upper() if h else '' for h in table[0]]
                    
                    date_col = None
                    desc_col = None
                    value_col = None
                    debit_col = None
                    credit_col = None
                    type_col = None
                    
                    for i, h in enumerate(header):
                        if any(x in h for x in ['DATA', 'DATE', 'DT']):
                            date_col = i
                        elif any(x in h for x in ['DESCRIÇÃO', 'DESCRICAO', 'HISTÓRICO', 'HISTORICO', 'LANÇAMENTO', 'LANCAMENTO', 'MEMO']):
                            desc_col = i
                        elif any(x in h for x in ['DÉBITO', 'DEBITO', 'SAÍDA', 'SAIDA', 'DÉB', 'DEB']):
                            debit_col = i
                        elif any(x in h for x in ['CRÉDITO', 'CREDITO', 'ENTRADA', 'CRÉD', 'CRED']):
                            credit_col = i
                        elif any(x in h for x in ['VALOR', 'VALUE', 'QUANTIA', 'MONTANTE']):
                            value_col = i
                        elif any(x in h for x in ['TIPO', 'TYPE', 'D/C', 'C/D']):
                            type_col = i
                    
                    # Processar linhas da tabela
                    for row in table[1:]:
                        if not row or all(not cell for cell in row):
                            continue
                        
                        # Extrair data
                        date_val = None
                        if date_col is not None and date_col < len(row) and row[date_col]:
                            date_str = str(row[date_col]).strip()
                            date_match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', date_str)
                            if date_match:
                                date_val = date_match.group(1)
                                if len(date_val) == 5:
                                    date_val += "/2026"
                        
                        if not date_val:
                            continue
                        
                        # Extrair descrição
                        description = ""
                        if desc_col is not None and desc_col < len(row) and row[desc_col]:
                            description = str(row[desc_col]).strip()
                        
                        # Extrair valor - tratar colunas separadas de débito/crédito
                        amount = 0
                        trans_type = None
                        
                        if debit_col is not None and credit_col is not None:
                            # Colunas separadas de débito e crédito
                            debit_val = row[debit_col] if debit_col < len(row) else None
                            credit_val = row[credit_col] if credit_col < len(row) else None
                            
                            if debit_val and str(debit_val).strip():
                                amount = parse_brazilian_number(str(debit_val))
                                if amount != 0:
                                    amount = abs(amount)  # Débito é sempre positivo no registro, mas representa saída
                                    trans_type = 'D'
                            
                            if credit_val and str(credit_val).strip() and trans_type is None:
                                amount = parse_brazilian_number(str(credit_val))
                                if amount != 0:
                                    amount = abs(amount)
                                    trans_type = 'C'
                        
                        elif value_col is not None and value_col < len(row) and row[value_col]:
                            # Coluna única de valor
                            amount = parse_brazilian_number(str(row[value_col]))
                            
                            # Verificar coluna de tipo se existir
                            if type_col is not None and type_col < len(row) and row[type_col]:
                                type_str = str(row[type_col]).upper().strip()
                                if 'D' in type_str or 'DÉB' in type_str or 'SAÍ' in type_str:
                                    trans_type = 'D'
                                    amount = abs(amount)
                                elif 'C' in type_str or 'CRÉ' in type_str or 'ENT' in type_str:
                                    trans_type = 'C'
                                    amount = abs(amount)
                            
                            if trans_type is None:
                                trans_type = 'C' if amount > 0 else 'D'
                                amount = abs(amount)
                        
                        if description and amount != 0 and trans_type:
                            transactions.append({
                                'date': date_val,
                                'description': description,
                                'amount': amount,
                                'transaction_type': trans_type
                            })
            
            # Se não encontrou tabelas, tentar extração por texto
            if not transactions:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"
                
                lines = full_text.split('\n')
                
                for line in lines:
                    # Tentar encontrar data
                    date_match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', line)
                    if not date_match:
                        continue
                    
                    date_str = date_match.group(1)
                    if len(date_str) == 5:
                        date_str += "/2025"
                    
                    # Buscar múltiplos valores na linha
                    # Formato brasileiro: 1.234,56 ou -1.234,56 ou (1.234,56) ou 1.234,56- (Santander)
                    value_matches = re.findall(r'([+-]?\s*\d{1,3}(?:\.\d{3})*,\d{2}-?|\(\d{1,3}(?:\.\d{3})*,\d{2}\))', line)
                    
                    if not value_matches:
                        continue
                    
                    # Pegar o último valor encontrado (geralmente é o valor da transação)
                    value_str = value_matches[-1]
                    
                    # Detectar se é negativo: parênteses, sinal no início OU no final (Santander)
                    is_negative = value_str.startswith('(') or value_str.startswith('-') or value_str.endswith('-')
                    value_str = value_str.replace('(', '').replace(')', '').replace('.', '').replace(',', '.').replace(' ', '').replace('-', '').replace('+', '')
                    
                    try:
                        amount = float(value_str)
                        if is_negative:
                            amount = -amount
                    except:
                        continue
                    
                    # Extrair descrição
                    date_end = date_match.end()
                    value_start = line.find(value_matches[-1])
                    description = line[date_end:value_start].strip() if value_start > date_end else ""
                    description = re.sub(r'\s+', ' ', description)
                    
                    # Ignorar linhas de cabeçalho/resumo
                    if any(x in description.lower() for x in ['saldo', 'total', 'anterior', 'entradas', 'saídas']):
                        continue
                    
                    # Verificar indicadores de débito/crédito na linha
                    line_upper = line.upper()
                    trans_type = None
                    
                    if any(x in line_upper for x in ['DÉBITO', 'DEBITO', 'DÉB', 'DEB', 'SAÍDA', 'SAIDA', 'PAGAMENTO', 'TRANSF ENV', 'PIX ENV']):
                        trans_type = 'D'
                        amount = abs(amount)
                    elif any(x in line_upper for x in ['CRÉDITO', 'CREDITO', 'CRÉD', 'CRED', 'ENTRADA', 'RECEBIDO', 'TRANSF REC', 'PIX REC']):
                        trans_type = 'C'
                        amount = abs(amount)
                    else:
                        # Usar o sinal do valor para determinar o tipo
                        trans_type = 'C' if amount > 0 else 'D'
                        amount = abs(amount)
                    
                    if description and amount != 0:
                        transactions.append({
                            'date': date_str,
                            'description': description,
                            'amount': amount,
                            'transaction_type': trans_type
                        })
        
        if not transactions:
            raise HTTPException(status_code=400, detail="Não foi possível extrair transações do PDF. Verifique se o formato é compatível.")
        
        return transactions
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar PDF: {str(e)}")

def parse_santander_format(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Parser especializado para extratos do Santander"""
    transactions = []
    
    # No formato Santander:
    # - Coluna ~5 (Unnamed: 5): Data
    # - Coluna ~6 (Unnamed: 6): Descrição
    # - Coluna ~9 (Unnamed: 9): Entradas (créditos) - valores float
    # - Coluna ~14 (Unnamed: 14): Saídas (débitos) - valores com "-" no final
    
    # Identificar colunas por conteúdo
    date_col = None
    desc_col = None
    credit_col = None
    debit_col = None
    
    # Primeiro, encontrar coluna de data
    for i in range(len(df.columns)):
        col_data = df.iloc[:50, i].dropna()
        date_count = 0
        for val in col_data:
            if re.match(r'^\d{2}/\d{2}(/\d{4})?$', str(val).strip()):
                date_count += 1
        if date_count >= 3:
            date_col = i
            logger.info(f"Santander: Coluna de data detectada no índice {i}")
            break
    
    # Encontrar coluna de débitos (valores terminando em "-")
    for i in range(len(df.columns)):
        col_data = df.iloc[:50, i].dropna()
        neg_count = 0
        for val in col_data:
            val_str = str(val).strip()
            if val_str.endswith('-') and re.search(r'\d', val_str):
                neg_count += 1
        if neg_count >= 5:
            debit_col = i
            logger.info(f"Santander: Coluna de débitos detectada no índice {i}")
            break
    
    # Encontrar coluna de créditos (valores numéricos positivos, tipo float)
    for i in range(len(df.columns)):
        if i == debit_col or i == date_col:
            continue
        col_data = df.iloc[:50, i].dropna()
        num_count = 0
        for val in col_data:
            # Valores float positivos indicam créditos
            if isinstance(val, (int, float)) and val > 0:
                num_count += 1
        if num_count >= 5:
            credit_col = i
            logger.info(f"Santander: Coluna de créditos detectada no índice {i} (valores numéricos)")
            break
    
    # Encontrar coluna de descrição
    if date_col is not None:
        for i in [date_col + 1, date_col + 2, date_col - 1]:
            if 0 <= i < len(df.columns) and i != credit_col and i != debit_col:
                col_data = df.iloc[:30, i].dropna()
                text_count = 0
                for val in col_data:
                    if isinstance(val, str) and len(val) > 5 and not re.match(r'^[\d.,\-+\s]+$', val):
                        text_count += 1
                if text_count >= 3:
                    desc_col = i
                    logger.info(f"Santander: Coluna de descrição detectada no índice {i}")
                    break
    
    logger.info(f"Santander Parser - date_col: {date_col}, desc_col: {desc_col}, credit_col: {credit_col}, debit_col: {debit_col}")
    
    # Variável para manter a última data válida
    last_date = None
    
    # Processar linhas
    for idx, row in df.iterrows():
        # Extrair data
        date_val = None
        if date_col is not None:
            raw_date = row.iloc[date_col]
            if pd.notna(raw_date):
                date_str = str(raw_date).strip()
                date_match = re.match(r'^(\d{2}/\d{2})(/\d{4})?$', date_str)
                if date_match:
                    date_val = date_match.group(1)
                    if date_match.group(2):
                        date_val += date_match.group(2)
                    else:
                        date_val += "/2025"
                    last_date = date_val
        
        # Usar última data válida se não tiver data na linha
        if date_val is None and last_date is not None:
            date_val = last_date
        
        if date_val is None:
            continue
        
        # Extrair descrição
        description = ""
        if desc_col is not None and pd.notna(row.iloc[desc_col]):
            description = str(row.iloc[desc_col]).strip()
        
        # Se descrição vazia ou é texto de cabeçalho/resumo, ignorar
        if not description:
            continue
        if 'saldo' in description.lower():
            continue
        if any(x in description.lower() for x in ['total', 'entradas', 'saídas', 'créditos', 'débitos', 'siglas', 'notas', 'bolsa']):
            continue
        
        # Extrair valores de crédito e débito
        amount = 0
        trans_type = None
        
        # Verificar CRÉDITO primeiro (valores numéricos float)
        if credit_col is not None and pd.notna(row.iloc[credit_col]):
            credit_val = row.iloc[credit_col]
            # Verificar se é número (não string de cabeçalho)
            if isinstance(credit_val, (int, float)) and credit_val > 0:
                amount = float(credit_val)
                trans_type = 'C'
            elif isinstance(credit_val, str):
                credit_str = str(credit_val).strip()
                if credit_str and not any(x in credit_str.lower() for x in ['entrada', 'crédito', 'credito', 'r$']):
                    parsed = parse_brazilian_number(credit_str)
                    if parsed > 0:
                        amount = parsed
                        trans_type = 'C'
        
        # Verificar DÉBITO (valores com "-" no final)
        if trans_type is None and debit_col is not None and pd.notna(row.iloc[debit_col]):
            debit_str = str(row.iloc[debit_col]).strip()
            if debit_str and not any(x in debit_str.lower() for x in ['saída', 'saida', 'débito', 'debito', 'r$']):
                if debit_str.endswith('-') or re.search(r'\d', debit_str):
                    debit_val = parse_brazilian_number(debit_str)
                    if debit_val != 0:
                        amount = abs(debit_val)
                        trans_type = 'D'
        
        # Adicionar transação se válida
        if description and amount > 0 and trans_type:
            transactions.append({
                'date': date_val,
                'description': description,
                'amount': amount,
                'transaction_type': trans_type
            })
    
    if not transactions:
        raise HTTPException(status_code=400, detail="Não foi possível extrair transações do arquivo Santander.")
    
    # Contar por tipo
    credits = len([t for t in transactions if t['transaction_type'] == 'C'])
    debits = len([t for t in transactions if t['transaction_type'] == 'D'])
    logger.info(f"Santander Parser: {len(transactions)} transações extraídas ({credits} créditos, {debits} débitos)")
    
    return transactions

def convert_legacy_excel_with_ssconvert(file_content: bytes) -> bytes:
    """Converte arquivos Excel antigos (BIFF5/Excel 95) para XLSX usando ssconvert"""
    import subprocess
    import tempfile
    
    # Criar arquivos temporários
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as input_file:
        input_file.write(file_content)
        input_path = input_file.name
    
    output_path = input_path.replace('.xls', '_converted.xlsx')
    
    try:
        # Usar ssconvert para converter
        result = subprocess.run(
            ['ssconvert', input_path, output_path],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0 and os.path.exists(output_path):
            with open(output_path, 'rb') as f:
                return f.read()
        else:
            logger.warning(f"ssconvert falhou: {result.stderr}")
            return None
    except Exception as e:
        logger.warning(f"Erro ao usar ssconvert: {e}")
        return None
    finally:
        # Limpar arquivos temporários
        try:
            os.unlink(input_path)
            if os.path.exists(output_path):
                os.unlink(output_path)
        except:
            pass

def parse_excel_statement(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse Excel/CSV bank statement - Suporta múltiplos formatos incluindo BIFF5 (Excel 5.0/95)"""
    try:
        transactions = []
        df = None
        
        # Tentar ler como Excel primeiro, depois CSV
        try:
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        except Exception as e1:
            logger.debug(f"openpyxl falhou: {e1}")
            try:
                df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
            except Exception as e2:
                logger.debug(f"xlrd falhou: {e2}")
                # Tentar converter arquivo legado (BIFF5/Excel 95) usando ssconvert
                logger.info("Tentando converter arquivo Excel legado com ssconvert...")
                converted_content = convert_legacy_excel_with_ssconvert(file_content)
                if converted_content:
                    try:
                        df = pd.read_excel(io.BytesIO(converted_content), engine='openpyxl')
                        logger.info("Arquivo Excel legado convertido com sucesso!")
                    except Exception as e3:
                        logger.warning(f"Falha ao ler arquivo convertido: {e3}")
        
        if df is None or df.empty:
            # Tentar como CSV com diferentes encodings
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding, sep=None, engine='python')
                    if not df.empty:
                        break
                except:
                    continue
        
        if df is None or df.empty:
            raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo. Verifique o formato. Para arquivos Excel muito antigos (Excel 5.0/95), tente salvar em formato mais recente.")
        
        # Limpar dados
        df = df.dropna(how='all')
        
        # =====================================================
        # DETECÇÃO DE FORMATO SANTANDER (colunas Unnamed)
        # =====================================================
        unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col) or 'unnamed' in str(col).lower()]
        
        if len(unnamed_cols) > 10:
            logger.info("Detectado formato Santander - processando com parser especializado")
            return parse_santander_format(df)
        
        # =====================================================
        # FORMATO PADRÃO COM COLUNAS NOMEADAS
        # =====================================================
        
        # Normalizar nomes de colunas
        df.columns = [str(col).strip().upper() for col in df.columns]
        
        # Mapear colunas por padrões conhecidos
        col_mapping = {
            'date': None,
            'description': None,
            'value': None,
            'debit': None,
            'credit': None,
            'type': None
        }
        
        for col in df.columns:
            col_lower = col.lower()
            if any(x in col_lower for x in ['data', 'date', 'dt', 'dia']):
                col_mapping['date'] = col
            elif any(x in col_lower for x in ['descrição', 'descricao', 'histórico', 'historico', 'memo', 'lançamento', 'lancamento', 'descriç']):
                col_mapping['description'] = col
            elif any(x in col_lower for x in ['débito', 'debito', 'déb', 'deb', 'saída', 'saida']):
                col_mapping['debit'] = col
            elif any(x in col_lower for x in ['crédito', 'credito', 'créd', 'cred', 'entrada']):
                col_mapping['credit'] = col
            elif any(x in col_lower for x in ['valor', 'value', 'quantia', 'montante', 'vlr']):
                col_mapping['value'] = col
            elif any(x in col_lower for x in ['tipo', 'type', 'd/c', 'c/d', 'natureza']):
                col_mapping['type'] = col
        
        # Se não encontrou colunas pelos nomes, tentar detectar automaticamente
        if not col_mapping['date']:
            for col in df.columns:
                sample = df[col].dropna().head(5)
                for val in sample:
                    if isinstance(val, datetime):
                        col_mapping['date'] = col
                        break
                    if isinstance(val, str) and re.search(r'\d{2}/\d{2}', str(val)):
                        col_mapping['date'] = col
                        break
                if col_mapping['date']:
                    break
        
        if not col_mapping['description']:
            for col in df.columns:
                if col == col_mapping['date']:
                    continue
                sample = df[col].dropna().head(5)
                if all(isinstance(v, str) and len(str(v)) > 5 for v in sample):
                    col_mapping['description'] = col
                    break
        
        # Processar cada linha
        for _, row in df.iterrows():
            # Extrair data
            date_val = None
            if col_mapping['date'] and pd.notna(row.get(col_mapping['date'])):
                raw_date = row[col_mapping['date']]
                if isinstance(raw_date, datetime):
                    date_val = raw_date.strftime('%d/%m/%Y')
                elif isinstance(raw_date, str):
                    date_match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', str(raw_date))
                    if date_match:
                        date_val = date_match.group(1)
                        if len(date_val) == 5:
                            date_val += "/2026"
            
            if not date_val:
                continue
            
            # Extrair descrição
            description = ""
            if col_mapping['description'] and pd.notna(row.get(col_mapping['description'])):
                description = str(row[col_mapping['description']]).strip()
            
            if not description:
                # Tentar encontrar qualquer texto longo na linha
                for col in df.columns:
                    val = row.get(col)
                    if pd.notna(val) and isinstance(val, str) and len(str(val)) > 10:
                        if val != date_val:
                            description = str(val).strip()
                            break
            
            # Extrair valor - IMPORTANTE: tratar colunas separadas de débito/crédito
            amount = 0
            trans_type = None
            
            # Caso 1: Colunas separadas de débito e crédito
            if col_mapping['debit'] is not None or col_mapping['credit'] is not None:
                debit_val = row.get(col_mapping['debit']) if col_mapping['debit'] else None
                credit_val = row.get(col_mapping['credit']) if col_mapping['credit'] else None
                
                # Verificar débito
                if pd.notna(debit_val):
                    parsed = parse_brazilian_number(str(debit_val))
                    if parsed != 0:
                        amount = abs(parsed)
                        trans_type = 'D'
                
                # Verificar crédito (só se não tiver débito)
                if trans_type is None and pd.notna(credit_val):
                    parsed = parse_brazilian_number(str(credit_val))
                    if parsed != 0:
                        amount = abs(parsed)
                        trans_type = 'C'
            
            # Caso 2: Coluna única de valor
            if trans_type is None and col_mapping['value']:
                val = row.get(col_mapping['value'])
                if pd.notna(val):
                    amount = parse_brazilian_number(str(val))
                    
                    # Verificar coluna de tipo
                    if col_mapping['type'] and pd.notna(row.get(col_mapping['type'])):
                        type_str = str(row[col_mapping['type']]).upper().strip()
                        if any(x in type_str for x in ['D', 'DÉB', 'DEB', 'SAÍ', 'SAI']):
                            trans_type = 'D'
                            amount = abs(amount)
                        elif any(x in type_str for x in ['C', 'CRÉ', 'CRE', 'ENT']):
                            trans_type = 'C'
                            amount = abs(amount)
                    
                    if trans_type is None:
                        trans_type = 'C' if amount > 0 else 'D'
                        amount = abs(amount)
            
            # Caso 3: Tentar encontrar valor em qualquer coluna numérica
            if trans_type is None:
                for col in df.columns:
                    if col in [col_mapping['date'], col_mapping['description']]:
                        continue
                    val = row.get(col)
                    if pd.notna(val):
                        try:
                            if isinstance(val, (int, float)) and val != 0:
                                amount = float(val)
                                trans_type = 'C' if amount > 0 else 'D'
                                amount = abs(amount)
                                break
                            elif isinstance(val, str):
                                parsed = parse_brazilian_number(val)
                                if parsed != 0:
                                    amount = parsed
                                    trans_type = 'C' if amount > 0 else 'D'
                                    amount = abs(amount)
                                    break
                        except:
                            pass
            
            # Adicionar transação se tiver dados válidos
            if description and amount != 0 and trans_type:
                transactions.append({
                    'date': date_val,
                    'description': description,
                    'amount': amount,
                    'transaction_type': trans_type
                })
        
        if not transactions:
            raise HTTPException(status_code=400, detail="Não foi possível extrair transações do arquivo. Verifique se o formato está correto.")
        
        return transactions
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")

def parse_ofx_statement(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse OFX bank statement"""
    try:
        ofx = ofxparse.OfxParser.parse(io.BytesIO(file_content))
        transactions = []
        
        for account in ofx.accounts:
            for trans in account.statement.transactions:
                amount = float(trans.amount)
                transactions.append({
                    'date': trans.date.strftime('%d/%m/%Y'),
                    'description': trans.memo or trans.payee or 'Sem descrição',
                    'document': trans.id or None,
                    'amount': amount,
                    'transaction_type': 'C' if amount > 0 else 'D'
                })
        
        return transactions
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar OFX: {str(e)}")

def calculate_similarity(str1: str, str2: str) -> float:
    """Calcula similaridade entre duas strings (0 a 1)"""
    str1 = str1.upper().strip()
    str2 = str2.upper().strip()
    
    # Exact match
    if str1 == str2:
        return 1.0
    
    # Substring match
    if str1 in str2 or str2 in str1:
        return 0.8
    
    # Word overlap
    words1 = set(str1.split())
    words2 = set(str2.split())
    
    if not words1 or not words2:
        return 0.0
    
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    
    return len(intersection) / len(union) if union else 0.0

async def classify_transaction(description: str, amount: float, trans_type: str, chart_id: str, company_id: str) -> Dict[str, Optional[str]]:
    """Classificar transação com aprendizado inteligente baseado em histórico"""
    
    description_upper = description.upper()
    
    # PASSO 1: Buscar no histórico de classificações da empresa
    history_items = await db.classification_history.find(
        {"company_id": company_id},
        {"_id": 0}
    ).sort("usage_count", -1).to_list(500)
    
    # Encontrar melhor match por similaridade
    best_match = None
    best_score = 0.0
    
    for item in history_items:
        # Só considerar mesmo tipo de transação
        if item['transaction_type'] == trans_type:
            score = calculate_similarity(description, item['description_pattern'])
            if score > best_score and score >= 0.6:  # Mínimo 60% similaridade
                best_score = score
                best_match = item
    
    if best_match:
        # Atualizar contador de uso
        await db.classification_history.update_one(
            {"id": best_match['id']},
            {
                "$inc": {"usage_count": 1},
                "$set": {"last_used": datetime.now(timezone.utc)}
            }
        )
        
        return {
            'debit_account': best_match['debit_account'],
            'credit_account': best_match['credit_account'],
            'status': 'CLASSIFICADO',
            'confidence': best_score
        }
    
    # PASSO 2: Buscar contas do plano (se existir)
    accounts = {}
    account_items = await db.account_items.find({"chart_id": chart_id}, {"_id": 0}).to_list(1000)
    
    accounts_by_code = {item['code']: item for item in account_items}
    accounts_by_type = defaultdict(list)
    for item in account_items:
        accounts_by_type[item['account_type']].append(item)
    
    # PASSO 3: Buscar regras configuradas
    rules = await db.classification_rules.find({}, {"_id": 0}).sort("priority", -1).to_list(100)
    
    # Tentar aplicar regras
    for rule in rules:
        if rule['keyword'].upper() in description_upper:
            debit_code = rule.get('debit_account_code')
            credit_code = rule.get('credit_account_code')
            
            # Se regra tem códigos válidos
            if debit_code and credit_code:
                if not account_items or (debit_code in accounts_by_code and credit_code in accounts_by_code):
                    return {
                        'debit_account': debit_code,
                        'credit_account': credit_code,
                        'status': 'CLASSIFICADO',
                        'confidence': 1.0
                    }
    
    # PASSO 4: Lógica automática baseada no tipo (ENTRADA/SAÍDA)
    if account_items:
        if trans_type == 'C':  # ENTRADA - Débito: Banco, Crédito: Receita
            banco_conta = None
            for conta in accounts_by_type['ATIVO']:
                if 'BANCO' in conta['description'].upper() or 'CAIXA' in conta['description'].upper():
                    banco_conta = conta['code']
                    break
            
            receita_conta = None
            for conta in accounts_by_type['RECEITA']:
                if 'RECEITA' in conta['description'].upper() or 'VENDA' in conta['description'].upper():
                    receita_conta = conta['code']
                    break
            
            if banco_conta and receita_conta:
                return {
                    'debit_account': banco_conta,
                    'credit_account': receita_conta,
                    'status': 'CLASSIFICADO',
                    'confidence': 0.7
                }
        
        elif trans_type == 'D':  # SAÍDA - Crédito: Banco, Débito: Despesa
            banco_conta = None
            for conta in accounts_by_type['ATIVO']:
                if 'BANCO' in conta['description'].upper() or 'CAIXA' in conta['description'].upper():
                    banco_conta = conta['code']
                    break
            
            despesa_conta = None
            if accounts_by_type['DESPESA']:
                despesa_conta = accounts_by_type['DESPESA'][0]['code']
            
            if despesa_conta and banco_conta:
                return {
                    'debit_account': despesa_conta,
                    'credit_account': banco_conta,
                    'status': 'CLASSIFICADO',
                    'confidence': 0.5
                }
    
    return {
        'debit_account': None,
        'credit_account': None,
        'status': 'CLASSIFICAR MANUALMENTE',
        'confidence': 0.0
    }

# ============= ROUTES =============

@api_router.get("/")
async def root():
    return {"message": "API Agente Contábil - Sistema Domínio"}

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    """Retorna estatísticas operacionais do escritório contábil"""
    
    # Mês atual para comparação
    now = datetime.now(timezone.utc)
    current_month = f"{now.month:02d}/{now.year}"
    current_year_month = f"{now.year}{now.month:02d}"
    
    # Buscar todas as empresas
    companies = await db.companies.find({}, {"_id": 0}).to_list(1000)
    total_companies = len(companies)
    
    # Buscar todos os extratos
    statements = await db.bank_statements.find({}, {"_id": 0}).to_list(10000)
    
    # Buscar todas as transações pendentes de classificação
    pending_transactions = await db.transactions.count_documents({"status": "CLASSIFICAR MANUALMENTE"})
    total_transactions = await db.transactions.count_documents({})
    classified_transactions = total_transactions - pending_transactions
    
    # Agrupar extratos por empresa e encontrar o último mês processado
    company_status = {}
    for company in companies:
        company_id = company['id']
        company_statements = [s for s in statements if s.get('company_id') == company_id]
        
        if not company_statements:
            company_status[company_id] = {
                'company_name': company['name'],
                'cnpj': company.get('cnpj', ''),
                'last_period': None,
                'last_period_date': None,
                'months_behind': 999,  # Nunca teve extrato
                'status': 'SEM_EXTRATO',
                'pending_transactions': 0
            }
        else:
            # Encontrar o período mais recente
            def parse_period(period_str):
                """Converte MM/YYYY para YYYYMM para ordenação"""
                try:
                    parts = period_str.split('/')
                    if len(parts) == 2:
                        month, year = parts
                        return f"{year}{month.zfill(2)}"
                except:
                    pass
                return "000000"
            
            sorted_statements = sorted(company_statements, key=lambda s: parse_period(s.get('period', '')), reverse=True)
            last_statement = sorted_statements[0]
            last_period = last_statement.get('period', '')
            last_period_key = parse_period(last_period)
            
            # Calcular meses de atraso
            try:
                if len(last_period_key) == 6:
                    last_year = int(last_period_key[:4])
                    last_month = int(last_period_key[4:])
                    current_year = now.year
                    current_month_num = now.month
                    months_behind = (current_year - last_year) * 12 + (current_month_num - last_month)
                else:
                    months_behind = 999
            except:
                months_behind = 999
            
            # Contar transações pendentes desta empresa
            company_pending = 0
            for s in company_statements:
                trans_pending = await db.transactions.count_documents({
                    "statement_id": s['id'],
                    "status": "CLASSIFICAR MANUALMENTE"
                })
                company_pending += trans_pending
            
            # Determinar status
            if months_behind <= 0:
                status = 'EM_DIA'
            elif months_behind <= 2:
                status = 'ATRASADA'
            else:
                status = 'MUITO_ATRASADA'
            
            company_status[company_id] = {
                'company_name': company['name'],
                'cnpj': company.get('cnpj', ''),
                'last_period': last_period,
                'last_period_date': last_period_key,
                'months_behind': months_behind,
                'status': status,
                'pending_transactions': company_pending
            }
    
    # Calcular indicadores
    companies_up_to_date = len([c for c in company_status.values() if c['status'] == 'EM_DIA'])
    companies_behind = len([c for c in company_status.values() if c['status'] == 'ATRASADA'])
    companies_very_behind = len([c for c in company_status.values() if c['status'] == 'MUITO_ATRASADA'])
    companies_no_statement = len([c for c in company_status.values() if c['status'] == 'SEM_EXTRATO'])
    
    # Calcular meses contábeis pendentes (soma de meses em atraso de todas as empresas)
    total_months_pending = sum(c['months_behind'] for c in company_status.values() if c['months_behind'] < 999)
    
    # Ranking de empresas mais atrasadas
    most_behind = sorted(
        [c for c in company_status.values() if c['months_behind'] > 0 and c['months_behind'] < 999],
        key=lambda x: x['months_behind'],
        reverse=True
    )[:10]
    
    # Empresas com mais pendências de classificação
    most_pending = sorted(
        [c for c in company_status.values() if c['pending_transactions'] > 0],
        key=lambda x: x['pending_transactions'],
        reverse=True
    )[:10]
    
    # Empresas sem extrato do mês atual
    companies_without_current = [
        c for c in company_status.values()
        if c['months_behind'] > 0 or c['status'] == 'SEM_EXTRATO'
    ]
    
    # Lista completa de empresas com status
    company_list = sorted(
        list(company_status.values()),
        key=lambda x: (x['months_behind'] if x['months_behind'] < 999 else 9999, x['company_name'])
    )
    
    return {
        "summary": {
            "total_companies": total_companies,
            "companies_up_to_date": companies_up_to_date,
            "companies_behind": companies_behind,
            "companies_very_behind": companies_very_behind,
            "companies_no_statement": companies_no_statement,
            "total_statements": len(statements),
            "total_transactions": total_transactions,
            "classified_transactions": classified_transactions,
            "pending_transactions": pending_transactions,
            "total_months_pending": total_months_pending
        },
        "most_behind_companies": most_behind,
        "most_pending_companies": most_pending,
        "companies_without_current_month": companies_without_current[:10],
        "all_companies_status": company_list
    }

# ============= ACCOUNTING PROCESSES =============

@api_router.get("/accounting-processes")
async def get_accounting_processes(
    company_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    status: Optional[str] = None,
    responsible: Optional[str] = None,
    is_archived: Optional[bool] = False
):
    """Lista processamentos contábeis com filtros"""
    query = {}
    
    if company_id:
        query["company_id"] = company_id
    if year:
        query["year"] = year
    if month:
        query["month"] = month
    if status:
        query["status"] = status
    if responsible:
        query["responsible"] = {"$regex": responsible, "$options": "i"}
    if is_archived is not None:
        query["is_archived"] = is_archived
    
    processes = await db.accounting_processes.find(query, {"_id": 0}).sort([
        ("year", -1),
        ("month", -1),
        ("company_name", 1)
    ]).to_list(1000)
    
    # Converter datas
    for p in processes:
        for date_field in ['created_at', 'updated_at', 'completed_at']:
            if isinstance(p.get(date_field), str):
                try:
                    p[date_field] = datetime.fromisoformat(p[date_field])
                except:
                    pass
    
    return processes

@api_router.get("/accounting-processes/grouped")
async def get_accounting_processes_grouped(
    company_id: Optional[str] = None,
    is_archived: bool = False
):
    """Retorna processamentos agrupados por empresa > ano > mês"""
    query = {"is_archived": is_archived}
    if company_id:
        query["company_id"] = company_id
    
    processes = await db.accounting_processes.find(query, {"_id": 0}).sort([
        ("company_name", 1),
        ("year", -1),
        ("month", -1)
    ]).to_list(10000)
    
    # Agrupar por empresa > ano > mês
    grouped = {}
    for p in processes:
        company_name = p['company_name']
        year = p['year']
        month = p['month']
        
        if company_name not in grouped:
            grouped[company_name] = {
                'company_id': p['company_id'],
                'company_name': company_name,
                'years': {}
            }
        
        if year not in grouped[company_name]['years']:
            grouped[company_name]['years'][year] = {'months': {}}
        
        grouped[company_name]['years'][year]['months'][month] = p
    
    return grouped

@api_router.get("/accounting-processes/stats")
async def get_accounting_processes_stats():
    """Retorna estatísticas dos processamentos"""
    all_processes = await db.accounting_processes.find({"is_archived": False}, {"_id": 0}).to_list(10000)
    
    stats = {
        'total': len(all_processes),
        'by_status': {},
        'companies_with_pending': set(),
        'overdue_count': 0,
        'in_progress_count': 0,
        'completed_count': 0
    }
    
    now = datetime.now(timezone.utc)
    current_year = now.year
    current_month = now.month
    
    for p in all_processes:
        # Contar por status
        status = p.get('status', 'NAO_INICIADO')
        stats['by_status'][status] = stats['by_status'].get(status, 0) + 1
        
        # Verificar se está atrasado (mês anterior ao atual e não concluído)
        p_year = p.get('year', 0)
        p_month = p.get('month', 0)
        is_past = (p_year < current_year) or (p_year == current_year and p_month < current_month)
        
        if status != 'CONCLUIDO' and is_past:
            stats['overdue_count'] += 1
            stats['companies_with_pending'].add(p.get('company_name', ''))
        
        if status == 'EM_PROCESSAMENTO':
            stats['in_progress_count'] += 1
        elif status == 'CONCLUIDO':
            stats['completed_count'] += 1
    
    stats['companies_with_pending'] = len(stats['companies_with_pending'])
    
    return stats

@api_router.post("/accounting-processes")
async def create_accounting_process(process: AccountingProcessCreate):
    """Cria um novo processamento contábil"""
    # Buscar nome da empresa
    company = await db.companies.find_one({"id": process.company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    # Verificar se já existe processamento para este período
    existing = await db.accounting_processes.find_one({
        "company_id": process.company_id,
        "year": process.year,
        "month": process.month,
        "is_archived": False
    })
    if existing:
        raise HTTPException(status_code=400, detail="Já existe um processamento para este período")
    
    process_obj = AccountingProcess(
        company_id=process.company_id,
        company_name=company['name'],
        year=process.year,
        month=process.month,
        period=f"{process.month:02d}/{process.year}",
        responsible=process.responsible,
        observations=process.observations
    )
    
    doc = process_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    if doc['completed_at']:
        doc['completed_at'] = doc['completed_at'].isoformat()
    
    await db.accounting_processes.insert_one(doc)
    return process_obj

@api_router.post("/accounting-processes/bulk-create")
async def bulk_create_accounting_processes(
    company_id: str,
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
    responsible: Optional[str] = None
):
    """Cria múltiplos processamentos de uma vez (para preencher meses pendentes)"""
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    created = []
    current = datetime(start_year, start_month, 1)
    end = datetime(end_year, end_month, 1)
    
    while current <= end:
        # Verificar se já existe
        existing = await db.accounting_processes.find_one({
            "company_id": company_id,
            "year": current.year,
            "month": current.month,
            "is_archived": False
        })
        
        if not existing:
            process_obj = AccountingProcess(
                company_id=company_id,
                company_name=company['name'],
                year=current.year,
                month=current.month,
                period=f"{current.month:02d}/{current.year}",
                responsible=responsible
            )
            
            doc = process_obj.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            doc['updated_at'] = doc['updated_at'].isoformat()
            if doc['completed_at']:
                doc['completed_at'] = doc['completed_at'].isoformat()
            
            await db.accounting_processes.insert_one(doc)
            created.append({"year": current.year, "month": current.month})
        
        # Próximo mês
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    
    return {"created": created, "count": len(created)}

@api_router.get("/accounting-processes/{process_id}")
async def get_accounting_process(process_id: str):
    """Retorna detalhes de um processamento"""
    process = await db.accounting_processes.find_one({"id": process_id}, {"_id": 0})
    if not process:
        raise HTTPException(status_code=404, detail="Processamento não encontrado")
    return process

@api_router.put("/accounting-processes/{process_id}")
async def update_accounting_process(process_id: str, update: AccountingProcessUpdate):
    """Atualiza um processamento"""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # Se status mudou para CONCLUIDO, registrar data de conclusão
    if update_data.get('status') == 'CONCLUIDO':
        update_data['completed_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.accounting_processes.update_one(
        {"id": process_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Processamento não encontrado")
    
    return await db.accounting_processes.find_one({"id": process_id}, {"_id": 0})

@api_router.delete("/accounting-processes/{process_id}")
async def delete_accounting_process(process_id: str):
    """Exclui um processamento"""
    result = await db.accounting_processes.delete_one({"id": process_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Processamento não encontrado")
    return {"message": "Processamento excluído"}

@api_router.post("/accounting-processes/{process_id}/archive")
async def archive_accounting_process(process_id: str):
    """Arquiva um processamento"""
    result = await db.accounting_processes.update_one(
        {"id": process_id},
        {"$set": {"is_archived": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Processamento não encontrado")
    return {"message": "Processamento arquivado"}

@api_router.post("/accounting-processes/archive-old")
async def archive_old_processes(months_old: int = 12):
    """Arquiva automaticamente processamentos concluídos há mais de X meses"""
    cutoff_date = datetime.now(timezone.utc)
    cutoff_year = cutoff_date.year
    cutoff_month = cutoff_date.month - months_old
    
    while cutoff_month <= 0:
        cutoff_year -= 1
        cutoff_month += 12
    
    # Buscar processamentos concluídos antigos
    query = {
        "status": "CONCLUIDO",
        "is_archived": False,
        "$or": [
            {"year": {"$lt": cutoff_year}},
            {"$and": [{"year": cutoff_year}, {"month": {"$lt": cutoff_month}}]}
        ]
    }
    
    result = await db.accounting_processes.update_many(
        query,
        {"$set": {"is_archived": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"archived_count": result.modified_count}

@api_router.post("/accounting-processes/{process_id}/link-statement")
async def link_statement_to_process(process_id: str, statement_id: str):
    """Vincula um extrato bancário a um processamento"""
    # Verificar se o processamento existe
    process = await db.accounting_processes.find_one({"id": process_id}, {"_id": 0})
    if not process:
        raise HTTPException(status_code=404, detail="Processamento não encontrado")
    
    # Verificar se o extrato existe
    statement = await db.bank_statements.find_one({"id": statement_id}, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    
    # Adicionar extrato ao processamento
    statement_ids = process.get('statement_ids', [])
    if statement_id not in statement_ids:
        statement_ids.append(statement_id)
    
    # Recalcular estatísticas
    total_trans = 0
    classified_trans = 0
    for sid in statement_ids:
        trans = await db.transactions.find({"statement_id": sid}, {"_id": 0}).to_list(10000)
        total_trans += len(trans)
        classified_trans += len([t for t in trans if t.get('status') == 'CLASSIFICADO'])
    
    await db.accounting_processes.update_one(
        {"id": process_id},
        {"$set": {
            "statement_ids": statement_ids,
            "total_transactions": total_trans,
            "classified_transactions": classified_trans,
            "pending_transactions": total_trans - classified_trans,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Extrato vinculado ao processamento"}

@api_router.get("/accounting-processes/responsibles/list")
async def get_responsibles():
    """Lista todos os responsáveis únicos"""
    processes = await db.accounting_processes.find(
        {"responsible": {"$ne": None, "$ne": ""}},
        {"responsible": 1, "_id": 0}
    ).to_list(10000)
    
    responsibles = list(set(p['responsible'] for p in processes if p.get('responsible')))
    return sorted(responsibles)

# ============= CONVERSOR OFX =============

class ExtractedTransaction(BaseModel):
    """Transação extraída de arquivo"""
    date: str
    description: str
    value: float
    type: str  # 'DEBIT' ou 'CREDIT'
    fit_id: str = Field(default_factory=lambda: str(uuid.uuid4())[:12])

class ConversionPreview(BaseModel):
    """Preview da conversão"""
    file_name: str
    file_type: str
    total_transactions: int
    total_credits: float
    total_debits: float
    balance: float
    transactions: List[ExtractedTransaction]

def extract_transactions_from_excel(file_content: bytes, file_name: str) -> List[Dict]:
    """Extrai transações de arquivo Excel/CSV - Suporta BIFF5 (Excel 5.0/95)"""
    transactions = []
    df = None
    
    # Tentar ler como Excel primeiro
    try:
        if file_name.lower().endswith('.csv'):
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding, sep=None, engine='python')
                    if not df.empty:
                        break
                except:
                    continue
        else:
            try:
                df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
            except Exception as e1:
                logger.debug(f"openpyxl falhou: {e1}")
                try:
                    df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
                except Exception as e2:
                    logger.debug(f"xlrd falhou: {e2}")
                    # Tentar converter arquivo legado (BIFF5/Excel 95) usando ssconvert
                    logger.info("Tentando converter arquivo Excel legado com ssconvert...")
                    converted_content = convert_legacy_excel_with_ssconvert(file_content)
                    if converted_content:
                        try:
                            df = pd.read_excel(io.BytesIO(converted_content), engine='openpyxl')
                            logger.info("Arquivo Excel legado convertido com sucesso!")
                        except Exception as e3:
                            logger.warning(f"Falha ao ler arquivo convertido: {e3}")
    except Exception as e:
        logger.error(f"Erro ao ler arquivo: {e}")
        return []
    
    if df is None or df.empty:
        return []
    
    # Remover linhas totalmente vazias
    df = df.dropna(how='all')
    
    # Detectar colunas por conteúdo
    col_indices = {
        'date': None,
        'description': None,
        'value': None,
        'debit': None,
        'credit': None
    }
    
    # Normalizar nomes de colunas
    col_names = [str(col).strip().upper() for col in df.columns]
    
    for i, col_name in enumerate(col_names):
        col_lower = col_name.lower()
        if any(x in col_lower for x in ['data', 'date', 'dt', 'dia']):
            col_indices['date'] = i
        elif any(x in col_lower for x in ['descrição', 'descricao', 'histórico', 'historico', 'memo', 'lançamento', 'lancamento']):
            col_indices['description'] = i
        elif any(x in col_lower for x in ['débito', 'debito', 'déb', 'deb', 'saída', 'saida']):
            col_indices['debit'] = i
        elif any(x in col_lower for x in ['crédito', 'credito', 'créd', 'cred', 'entrada']):
            col_indices['credit'] = i
        elif any(x in col_lower for x in ['valor', 'value', 'quantia', 'montante', 'vlr']):
            col_indices['value'] = i
    
    # Se não encontrou colunas pelos nomes, detectar automaticamente
    if col_indices['date'] is None:
        for i in range(len(df.columns)):
            sample = df.iloc[:20, i].dropna()
            date_count = 0
            for val in sample:
                if isinstance(val, datetime):
                    date_count += 1
                elif isinstance(val, str) and re.search(r'\d{2}/\d{2}', str(val)):
                    date_count += 1
            if date_count >= 3:
                col_indices['date'] = i
                break
    
    # Detectar coluna de débito por valores com "-" no final (formato Santander)
    if col_indices['debit'] is None:
        for i in range(len(df.columns)):
            sample = df.iloc[:50, i].dropna()
            neg_count = sum(1 for val in sample if str(val).strip().endswith('-'))
            if neg_count >= 3:
                col_indices['debit'] = i
                break
    
    # Processar linhas
    for idx, row in df.iterrows():
        try:
            # Extrair data
            date_val = None
            if col_indices['date'] is not None:
                raw_date = row.iloc[col_indices['date']]
                if pd.notna(raw_date):
                    if isinstance(raw_date, datetime):
                        date_val = raw_date.strftime('%d/%m/%Y')
                    elif isinstance(raw_date, str):
                        match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', str(raw_date))
                        if match:
                            date_val = match.group(1)
                            if len(date_val) == 5:
                                date_val += "/2025"
            
            if not date_val:
                continue
            
            # Extrair descrição
            description = ""
            if col_indices['description'] is not None and pd.notna(row.iloc[col_indices['description']]):
                description = str(row.iloc[col_indices['description']]).strip()
            
            if not description:
                # Tentar encontrar texto em outras colunas
                for i in range(len(row)):
                    if i != col_indices['date'] and i != col_indices['value'] and i != col_indices['debit'] and i != col_indices['credit']:
                        val = row.iloc[i]
                        if pd.notna(val) and isinstance(val, str) and len(val) > 5:
                            description = val.strip()
                            break
            
            if not description:
                continue
            
            # Ignorar linhas de cabeçalho/resumo
            desc_lower = description.lower()
            if any(x in desc_lower for x in ['saldo', 'total', 'anterior', 'entradas', 'saídas']):
                continue
            
            # Extrair valor
            value = 0
            
            # Caso 1: Colunas separadas de débito e crédito
            if col_indices['debit'] is not None or col_indices['credit'] is not None:
                debit_val = 0
                credit_val = 0
                
                if col_indices['debit'] is not None and pd.notna(row.iloc[col_indices['debit']]):
                    debit_str = str(row.iloc[col_indices['debit']]).strip()
                    if debit_str and not any(x in debit_str.lower() for x in ['saída', 'débito', 'r$']):
                        debit_val = abs(parse_brazilian_number(debit_str))
                
                if col_indices['credit'] is not None and pd.notna(row.iloc[col_indices['credit']]):
                    credit_raw = row.iloc[col_indices['credit']]
                    if isinstance(credit_raw, (int, float)) and credit_raw > 0:
                        credit_val = float(credit_raw)
                    elif isinstance(credit_raw, str):
                        credit_str = str(credit_raw).strip()
                        if credit_str and not any(x in credit_str.lower() for x in ['entrada', 'crédito', 'r$']):
                            credit_val = abs(parse_brazilian_number(credit_str))
                
                # valor = credito - debito
                if credit_val > 0:
                    value = credit_val
                elif debit_val > 0:
                    value = -debit_val
            
            # Caso 2: Coluna única de valor
            elif col_indices['value'] is not None and pd.notna(row.iloc[col_indices['value']]):
                value = parse_brazilian_number(str(row.iloc[col_indices['value']]))
            
            if value == 0:
                continue
            
            transactions.append({
                'date': date_val,
                'description': description,
                'value': value,
                'type': 'CREDIT' if value > 0 else 'DEBIT'
            })
            
        except Exception as e:
            logger.error(f"Erro ao processar linha {idx}: {e}")
            continue
    
    return transactions

def extract_transactions_from_pdf(file_content: bytes) -> List[Dict]:
    """Extrai transações de arquivo PDF"""
    transactions = []
    
    try:
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            # Primeiro, tentar extrair tabelas
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Detectar colunas
                    header = [str(h).upper() if h else '' for h in table[0]]
                    
                    date_col = None
                    desc_col = None
                    value_col = None
                    debit_col = None
                    credit_col = None
                    
                    for i, h in enumerate(header):
                        if any(x in h for x in ['DATA', 'DATE', 'DT']):
                            date_col = i
                        elif any(x in h for x in ['DESCRIÇÃO', 'DESCRICAO', 'HISTÓRICO', 'HISTORICO', 'MEMO']):
                            desc_col = i
                        elif any(x in h for x in ['DÉBITO', 'DEBITO', 'SAÍDA', 'SAIDA']):
                            debit_col = i
                        elif any(x in h for x in ['CRÉDITO', 'CREDITO', 'ENTRADA']):
                            credit_col = i
                        elif any(x in h for x in ['VALOR', 'VALUE']):
                            value_col = i
                    
                    for row in table[1:]:
                        if not row or all(not cell for cell in row):
                            continue
                        
                        # Data
                        date_val = None
                        if date_col is not None and date_col < len(row) and row[date_col]:
                            match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', str(row[date_col]))
                            if match:
                                date_val = match.group(1)
                                if len(date_val) == 5:
                                    date_val += "/2025"
                        
                        if not date_val:
                            continue
                        
                        # Descrição
                        description = ""
                        if desc_col is not None and desc_col < len(row) and row[desc_col]:
                            description = str(row[desc_col]).strip()
                        
                        if not description:
                            continue
                        
                        # Valor
                        value = 0
                        if debit_col is not None and credit_col is not None:
                            debit_val = row[debit_col] if debit_col < len(row) else None
                            credit_val = row[credit_col] if credit_col < len(row) else None
                            
                            if debit_val and str(debit_val).strip():
                                value = -abs(parse_brazilian_number(str(debit_val)))
                            elif credit_val and str(credit_val).strip():
                                value = abs(parse_brazilian_number(str(credit_val)))
                        elif value_col is not None and value_col < len(row) and row[value_col]:
                            value = parse_brazilian_number(str(row[value_col]))
                        
                        if value != 0:
                            transactions.append({
                                'date': date_val,
                                'description': description,
                                'value': value,
                                'type': 'CREDIT' if value > 0 else 'DEBIT'
                            })
            
            # Se não encontrou tabelas, extrair por texto
            if not transactions:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"
                
                for line in full_text.split('\n'):
                    date_match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{2}/\d{2})', line)
                    if not date_match:
                        continue
                    
                    date_str = date_match.group(1)
                    if len(date_str) == 5:
                        date_str += "/2025"
                    
                    # Buscar valores
                    value_matches = re.findall(r'([+-]?\s*\d{1,3}(?:\.\d{3})*,\d{2}-?)', line)
                    if not value_matches:
                        continue
                    
                    value_str = value_matches[-1]
                    is_negative = value_str.startswith('-') or value_str.endswith('-')
                    value_str = value_str.replace('.', '').replace(',', '.').replace('-', '').replace('+', '').strip()
                    
                    try:
                        value = float(value_str)
                        if is_negative:
                            value = -value
                    except:
                        continue
                    
                    # Descrição
                    date_end = date_match.end()
                    value_start = line.find(value_matches[-1])
                    description = line[date_end:value_start].strip() if value_start > date_end else ""
                    description = re.sub(r'\s+', ' ', description)
                    
                    if description and value != 0:
                        transactions.append({
                            'date': date_str,
                            'description': description,
                            'value': value,
                            'type': 'CREDIT' if value > 0 else 'DEBIT'
                        })
    
    except Exception as e:
        logger.error(f"Erro ao processar PDF: {e}")
    
    return transactions

def generate_ofx_content(transactions: List[Dict], bank_name: str = "BANCO") -> str:
    """Gera conteúdo OFX a partir das transações"""
    
    # Header OFX
    ofx_content = """OFXHEADER:100
DATA:OFXSGML
VERSION:102
SECURITY:NONE
ENCODING:USASCII
CHARSET:1252
COMPRESSION:NONE
OLDFILEUID:NONE
NEWFILEUID:NONE

<OFX>
<SIGNONMSGSRSV1>
<SONRS>
<STATUS>
<CODE>0
<SEVERITY>INFO
</STATUS>
<DTSERVER>{dtserver}
<LANGUAGE>POR
</SONRS>
</SIGNONMSGSRSV1>
<BANKMSGSRSV1>
<STMTTRNRS>
<TRNUID>1
<STATUS>
<CODE>0
<SEVERITY>INFO
</STATUS>
<STMTRS>
<CURDEF>BRL
<BANKACCTFROM>
<BANKID>0000
<ACCTID>0000000000
<ACCTTYPE>CHECKING
</BANKACCTFROM>
<BANKTRANLIST>
<DTSTART>{dtstart}
<DTEND>{dtend}
"""
    
    # Determinar datas
    now = datetime.now()
    dtserver = now.strftime('%Y%m%d%H%M%S')
    
    dates = []
    for t in transactions:
        try:
            parts = t['date'].split('/')
            if len(parts) == 3:
                day, month, year = parts
                dates.append(datetime(int(year), int(month), int(day)))
        except:
            pass
    
    if dates:
        dtstart = min(dates).strftime('%Y%m%d')
        dtend = max(dates).strftime('%Y%m%d')
    else:
        dtstart = dtend = now.strftime('%Y%m%d')
    
    ofx_content = ofx_content.format(
        dtserver=dtserver,
        dtstart=dtstart,
        dtend=dtend
    )
    
    # Adicionar transações
    for i, t in enumerate(transactions):
        # Converter data para formato OFX (YYYYMMDD)
        try:
            parts = t['date'].split('/')
            if len(parts) == 3:
                day, month, year = parts
                dt_posted = f"{year}{month.zfill(2)}{day.zfill(2)}"
            else:
                dt_posted = now.strftime('%Y%m%d')
        except:
            dt_posted = now.strftime('%Y%m%d')
        
        trn_type = t['type']
        trn_amt = t['value']
        fit_id = t.get('fit_id', str(uuid.uuid4())[:12])
        name = t['description'][:64]  # Limitar a 64 caracteres
        
        ofx_content += f"""<STMTTRN>
<TRNTYPE>{trn_type}
<DTPOSTED>{dt_posted}
<TRNAMT>{trn_amt:.2f}
<FITID>{fit_id}
<NAME>{name}
</STMTTRN>
"""
    
    # Footer OFX
    ofx_content += """</BANKTRANLIST>
<LEDGERBAL>
<BALAMT>0.00
<DTASOF>{dtasof}
</LEDGERBAL>
</STMTRS>
</STMTTRNRS>
</BANKMSGSRSV1>
</OFX>
""".format(dtasof=now.strftime('%Y%m%d'))
    
    return ofx_content

@api_router.post("/converter/preview")
async def preview_conversion(file: UploadFile = File(...)):
    """Faz preview das transações extraídas do arquivo"""
    
    file_content = await file.read()
    file_name = file.filename.lower()
    
    # Identificar tipo de arquivo
    if file_name.endswith('.pdf'):
        file_type = 'PDF'
        transactions = extract_transactions_from_pdf(file_content)
    elif file_name.endswith(('.xlsx', '.xls')):
        file_type = 'Excel'
        transactions = extract_transactions_from_excel(file_content, file_name)
    elif file_name.endswith('.csv'):
        file_type = 'CSV'
        transactions = extract_transactions_from_excel(file_content, file_name)
    else:
        raise HTTPException(status_code=400, detail="Formato de arquivo não suportado. Use PDF, XLSX, XLS ou CSV.")
    
    if not transactions:
        raise HTTPException(status_code=400, detail="Não foi possível extrair transações do arquivo. Verifique se o formato está correto.")
    
    # Calcular totais
    total_credits = sum(t['value'] for t in transactions if t['value'] > 0)
    total_debits = sum(abs(t['value']) for t in transactions if t['value'] < 0)
    balance = total_credits - total_debits
    
    # Criar objetos ExtractedTransaction
    extracted = [
        ExtractedTransaction(
            date=t['date'],
            description=t['description'],
            value=t['value'],
            type=t['type'],
            fit_id=str(uuid.uuid4())[:12]
        )
        for t in transactions
    ]
    
    return ConversionPreview(
        file_name=file.filename,
        file_type=file_type,
        total_transactions=len(transactions),
        total_credits=total_credits,
        total_debits=total_debits,
        balance=balance,
        transactions=extracted
    )

@api_router.post("/converter/generate-ofx")
async def generate_ofx_file(
    transactions: List[ExtractedTransaction],
    bank_name: str = "BANCO"
):
    """Gera arquivo OFX a partir das transações"""
    
    # Converter para dicionários
    trans_list = [
        {
            'date': t.date,
            'description': t.description,
            'value': t.value,
            'type': t.type,
            'fit_id': t.fit_id
        }
        for t in transactions
    ]
    
    # Gerar conteúdo OFX
    ofx_content = generate_ofx_content(trans_list, bank_name)
    
    # Salvar em arquivo temporário
    temp_file = f"/tmp/extrato_{uuid.uuid4()}.ofx"
    with open(temp_file, 'w', encoding='utf-8') as f:
        f.write(ofx_content)
    
    return FileResponse(
        temp_file,
        media_type='application/x-ofx',
        filename=f"extrato_convertido_{datetime.now().strftime('%Y%m%d_%H%M%S')}.ofx"
    )

@api_router.post("/converter/import-to-system")
async def import_converted_to_system(
    transactions: List[ExtractedTransaction],
    company_id: str,
    chart_id: str,
    bank_name: str = "BANCO",
    period: str = ""
):
    """Importa as transações convertidas diretamente para o sistema de conciliação"""
    
    # Verificar empresa e plano de contas
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    chart = await db.chart_of_accounts.find_one({"id": chart_id}, {"_id": 0})
    if not chart:
        raise HTTPException(status_code=404, detail="Plano de contas não encontrado")
    
    # Determinar período se não informado
    if not period:
        dates = [t.date for t in transactions]
        if dates:
            # Pegar o mês/ano mais comum
            parts = dates[0].split('/')
            if len(parts) == 3:
                period = f"{parts[1]}/{parts[2]}"
            else:
                period = datetime.now().strftime('%m/%Y')
        else:
            period = datetime.now().strftime('%m/%Y')
    
    # Calcular totais
    total_inflows = sum(t.value for t in transactions if t.value > 0)
    total_outflows = sum(abs(t.value) for t in transactions if t.value < 0)
    
    # Criar o bank statement
    statement_id = str(uuid.uuid4())
    statement = {
        'id': statement_id,
        'company_id': company_id,
        'chart_id': chart_id,
        'bank_name': bank_name,
        'period': period,
        'filename': f'convertido_ofx_{period.replace("/", "_")}.ofx',
        'total_transactions': len(transactions),
        'classified_count': 0,
        'manual_count': len(transactions),
        'total_inflows': total_inflows,
        'total_outflows': total_outflows,
        'balance': total_inflows - total_outflows,
        'status': 'COMPLETED',
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    
    await db.bank_statements.insert_one(statement)
    
    # Criar transações
    created_transactions = []
    for t in transactions:
        # Tentar classificar usando histórico e regras
        trans_type = 'C' if t.value > 0 else 'D'
        classification = await classify_transaction(
            t.description,
            abs(t.value),
            trans_type,
            chart_id,
            company_id
        )
        
        trans = {
            'id': str(uuid.uuid4()),
            'statement_id': statement_id,
            'date': t.date,
            'description': t.description,
            'amount': abs(t.value),
            'transaction_type': 'C' if t.value > 0 else 'D',
            'debit_account': classification['debit_account'],
            'credit_account': classification['credit_account'],
            'status': classification['status'],
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        
        await db.transactions.insert_one(trans)
        created_transactions.append(trans)
    
    # Atualizar contagem de classificados
    classified_count = len([t for t in created_transactions if t['status'] == 'CLASSIFICADO'])
    manual_count = len(transactions) - classified_count
    
    await db.bank_statements.update_one(
        {'id': statement_id},
        {'$set': {'classified_count': classified_count, 'manual_count': manual_count}}
    )
    
    return {
        'statement_id': statement_id,
        'total_transactions': len(transactions),
        'classified_count': classified_count,
        'manual_count': manual_count,
        'message': 'Extrato importado com sucesso para o sistema de conciliação'
    }

# Companies
@api_router.post("/companies", response_model=Company)
async def create_company(company: CompanyCreate):
    company_obj = Company(**company.model_dump())
    doc = company_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['cnpj'] = clean_cnpj(doc['cnpj'])
    await db.companies.insert_one(doc)
    return company_obj

@api_router.get("/companies", response_model=List[Company])
async def get_companies():
    companies = await db.companies.find({}, {"_id": 0}).to_list(1000)
    for c in companies:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return companies

@api_router.get("/companies/{company_id}", response_model=Company)
async def get_company(company_id: str):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    if isinstance(company.get('created_at'), str):
        company['created_at'] = datetime.fromisoformat(company['created_at'])
    return company

@api_router.put("/companies/{company_id}", response_model=Company)
async def update_company(company_id: str, company: CompanyCreate):
    update_data = company.model_dump()
    update_data['cnpj'] = clean_cnpj(update_data['cnpj'])
    result = await db.companies.update_one({"id": company_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return await get_company(company_id)

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str):
    result = await db.companies.delete_one({"id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return {"message": "Empresa excluída com sucesso"}

# Chart of Accounts
@api_router.post("/chart-of-accounts", response_model=ChartOfAccounts)
async def create_chart(chart: ChartOfAccountsCreate):
    chart_obj = ChartOfAccounts(**chart.model_dump())
    doc = chart_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.chart_of_accounts.insert_one(doc)
    return chart_obj

@api_router.get("/chart-of-accounts", response_model=List[ChartOfAccounts])
async def get_charts(company_id: Optional[str] = None):
    query = {"company_id": company_id} if company_id else {}
    charts = await db.chart_of_accounts.find(query, {"_id": 0}).to_list(1000)
    for c in charts:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return charts

@api_router.delete("/chart-of-accounts/{chart_id}")
async def delete_chart(chart_id: str):
    result = await db.chart_of_accounts.delete_one({"id": chart_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano de contas não encontrado")
    await db.account_items.delete_many({"chart_id": chart_id})
    return {"message": "Plano de contas excluído com sucesso"}

@api_router.post("/chart-of-accounts/{chart_id}/import")
async def import_chart_accounts(
    chart_id: str,
    file: UploadFile = File(...)
):
    """Importar plano de contas em massa via Excel
    Formato esperado: Código | Descrição | Classificação | Tipo
    """
    try:
        # Verificar se o plano existe
        chart = await db.chart_of_accounts.find_one({"id": chart_id}, {"_id": 0})
        if not chart:
            raise HTTPException(status_code=404, detail="Plano de contas não encontrado")
        
        # Ler arquivo Excel
        content = await file.read()
        
        # Tentar diferentes formatos
        df = None
        try:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
        except:
            try:
                df = pd.read_excel(io.BytesIO(content), engine='xlrd')
            except:
                raise HTTPException(status_code=400, detail="Erro ao ler arquivo. Use formato XLSX ou XLS válido")
        
        # Normalizar nomes de colunas
        df.columns = df.columns.str.strip()
        
        # Verificar formato: Código, Descrição, Classificação, Tipo
        required_columns = ['Código', 'Descrição', 'Classificação', 'Tipo']
        
        # Permitir variações de nomenclatura
        column_mapping = {
            'codigo': 'Código',
            'código': 'Código',
            'descricao': 'Descrição',
            'descrição': 'Descrição',
            'classificacao': 'Classificação',
            'classificação': 'Classificação',
            'tipo': 'Tipo'
        }
        
        # Mapear colunas
        df.columns = [column_mapping.get(col.lower(), col) for col in df.columns]
        
        # Validar se todas as colunas obrigatórias existem
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Arquivo deve conter as colunas: Código, Descrição, Classificação, Tipo. Faltando: {', '.join(missing_cols)}"
            )
        
        # Processar e inserir contas
        imported_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Validar dados
                codigo_seq = str(row['Código']).strip() if pd.notna(row['Código']) else ''
                descricao = str(row['Descrição']).strip() if pd.notna(row['Descrição']) else ''
                classificacao = str(row['Classificação']).strip() if pd.notna(row['Classificação']) else ''
                tipo = str(row['Tipo']).upper().strip() if pd.notna(row['Tipo']) else ''
                
                # Validações
                if not classificacao or classificacao == 'nan':
                    errors.append(f"Linha {idx + 2}: Classificação vazia")
                    continue
                
                if not descricao or descricao == 'nan':
                    errors.append(f"Linha {idx + 2}: Descrição vazia")
                    continue
                
                if tipo not in ['ATIVO', 'PASSIVO', 'RECEITA', 'DESPESA']:
                    errors.append(f"Linha {idx + 2}: Tipo '{tipo}' inválido. Use: ATIVO, PASSIVO, RECEITA ou DESPESA")
                    continue
                
                # Usar a Classificação como código da conta (ex: 1.1.1.01.0001)
                codigo_conta = classificacao
                
                # Verificar se conta já existe
                existing = await db.account_items.find_one({
                    "chart_id": chart_id,
                    "code": codigo_conta
                }, {"_id": 0})
                
                if existing:
                    # Atualizar conta existente
                    await db.account_items.update_one(
                        {"chart_id": chart_id, "code": codigo_conta},
                        {"$set": {
                            "description": descricao,
                            "account_type": tipo
                        }}
                    )
                else:
                    # Criar nova conta
                    account = AccountItem(
                        chart_id=chart_id,
                        code=codigo_conta,
                        description=descricao,
                        account_type=tipo
                    )
                    doc = account.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.account_items.insert_one(doc)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Linha {idx + 2}: {str(e)}")
        
        return {
            "message": f"Importação concluída: {imported_count} contas processadas",
            "imported_count": imported_count,
            "errors": errors if errors else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")

# Account Items
@api_router.post("/account-items", response_model=AccountItem)
async def create_account_item(item: AccountItemCreate):
    item_obj = AccountItem(**item.model_dump())
    doc = item_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.account_items.insert_one(doc)
    return item_obj

@api_router.get("/account-items", response_model=List[AccountItem])
async def get_account_items(chart_id: Optional[str] = None):
    query = {"chart_id": chart_id} if chart_id else {}
    items = await db.account_items.find(query, {"_id": 0}).to_list(1000)
    for item in items:
        if isinstance(item.get('created_at'), str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
    return items

@api_router.delete("/account-items/{item_id}")
async def delete_account_item(item_id: str):
    result = await db.account_items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    return {"message": "Conta excluída com sucesso"}

# Classification Rules
@api_router.post("/classification-rules", response_model=ClassificationRule)
async def create_rule(rule: ClassificationRuleCreate):
    rule_obj = ClassificationRule(**rule.model_dump())
    doc = rule_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.classification_rules.insert_one(doc)
    return rule_obj

@api_router.get("/classification-rules", response_model=List[ClassificationRule])
async def get_rules():
    rules = await db.classification_rules.find({}, {"_id": 0}).to_list(1000)
    for r in rules:
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return rules

@api_router.delete("/classification-rules/{rule_id}")
async def delete_rule(rule_id: str):
    result = await db.classification_rules.delete_one({"id": rule_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Regra não encontrada")
    return {"message": "Regra excluída com sucesso"}

# Classification History - Memória Inteligente
@api_router.get("/classification-history")
async def get_classification_history(company_id: Optional[str] = None):
    """Retorna o histórico de classificações aprendidas"""
    query = {"company_id": company_id} if company_id else {}
    history = await db.classification_history.find(query, {"_id": 0}).sort("usage_count", -1).to_list(1000)
    return history

@api_router.delete("/classification-history/{history_id}")
async def delete_classification_history(history_id: str):
    """Remove um registro do histórico de aprendizado"""
    result = await db.classification_history.delete_one({"id": history_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado")
    return {"message": "Registro removido do histórico de aprendizado"}

@api_router.delete("/classification-history/company/{company_id}")
async def clear_company_history(company_id: str):
    """Limpa todo o histórico de aprendizado de uma empresa"""
    result = await db.classification_history.delete_many({"company_id": company_id})
    return {"message": f"{result.deleted_count} registros removidos"}

@api_router.put("/classification-rules/{rule_id}", response_model=ClassificationRule)
async def update_rule(rule_id: str, rule: ClassificationRuleCreate):
    rule_data = rule.model_dump()
    result = await db.classification_rules.update_one(
        {"id": rule_id},
        {"$set": rule_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Regra não encontrada")
    
    updated_rule = await db.classification_rules.find_one({"id": rule_id}, {"_id": 0})
    if isinstance(updated_rule.get('created_at'), str):
        updated_rule['created_at'] = datetime.fromisoformat(updated_rule['created_at'])
    return updated_rule

# Bank Statements - Upload and Process
@api_router.post("/bank-statements/upload")
async def upload_statement(
    file: UploadFile = File(...),
    company_id: str = Query(...),
    chart_id: str = Query(...),
    bank_name: str = Query(...),
    period: str = Query(...)
):
    # Ler arquivo
    content = await file.read()
    
    # Detectar tipo e parsear
    transactions = []
    if file.filename.endswith('.ofx'):
        transactions = parse_ofx_statement(content)
    elif file.filename.endswith('.pdf'):
        transactions = parse_pdf_statement(content)
    elif file.filename.endswith(('.xlsx', '.xls', '.csv')):
        transactions = parse_excel_statement(content)
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Use OFX, PDF, Excel ou CSV")
    
    # Criar statement
    statement_id = str(uuid.uuid4())
    
    # Classificar transações
    classified_transactions = []
    for trans in transactions:
        classification = await classify_transaction(
            trans['description'],
            trans['amount'],
            trans['transaction_type'],
            chart_id,
            company_id  # Adicionar company_id
        )
        
        trans_obj = Transaction(
            id=str(uuid.uuid4()),
            statement_id=statement_id,
            date=trans['date'],
            description=trans['description'],
            document=trans.get('document'),
            amount=abs(trans['amount']),
            transaction_type=trans['transaction_type'],
            debit_account=classification['debit_account'],
            credit_account=classification['credit_account'],
            status=classification['status']
        )
        classified_transactions.append(trans_obj)
    
    # Calcular estatísticas
    total_inflows = sum(t.amount for t in classified_transactions if t.transaction_type == 'C')
    total_outflows = sum(t.amount for t in classified_transactions if t.transaction_type == 'D')
    classified_count = sum(1 for t in classified_transactions if t.status == 'CLASSIFICADO')
    manual_count = len(classified_transactions) - classified_count
    
    # Criar statement
    statement = BankStatement(
        id=statement_id,
        company_id=company_id,
        chart_id=chart_id,
        filename=file.filename,
        bank_name=bank_name,
        period=period,
        total_transactions=len(classified_transactions),
        classified_count=classified_count,
        manual_count=manual_count,
        total_inflows=total_inflows,
        total_outflows=total_outflows,
        balance=total_inflows - total_outflows,
        status='COMPLETED',
        processed_at=datetime.now(timezone.utc)
    )
    
    # Salvar no banco
    statement_doc = statement.model_dump()
    statement_doc['created_at'] = statement_doc['created_at'].isoformat()
    statement_doc['processed_at'] = statement_doc['processed_at'].isoformat()
    await db.bank_statements.insert_one(statement_doc)
    
    for trans in classified_transactions:
        await db.transactions.insert_one(trans.model_dump())
    
    return {
        "statement": statement,
        "transactions": classified_transactions
    }

@api_router.get("/bank-statements", response_model=List[BankStatement])
async def get_statements(company_id: Optional[str] = None):
    query = {"company_id": company_id} if company_id else {}
    statements = await db.bank_statements.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for s in statements:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
        if isinstance(s.get('processed_at'), str):
            s['processed_at'] = datetime.fromisoformat(s['processed_at'])
    return statements

@api_router.get("/bank-statements/{statement_id}", response_model=BankStatement)
async def get_statement(statement_id: str):
    statement = await db.bank_statements.find_one({"id": statement_id}, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    if isinstance(statement.get('created_at'), str):
        statement['created_at'] = datetime.fromisoformat(statement['created_at'])
    if isinstance(statement.get('processed_at'), str):
        statement['processed_at'] = datetime.fromisoformat(statement['processed_at'])
    return statement

@api_router.get("/bank-statements/{statement_id}/transactions", response_model=List[Transaction])
async def get_transactions(statement_id: str):
    transactions = await db.transactions.find({"statement_id": statement_id}, {"_id": 0}).to_list(1000)
    return transactions

# Modelo para atualização em massa - DEVE VIR ANTES do endpoint com path parameter
class BulkUpdateRequest(BaseModel):
    transaction_ids: List[str]
    update_data: dict

@api_router.put("/transactions/bulk-update")
async def bulk_update_transactions(request: BulkUpdateRequest):
    """Atualiza múltiplas transações de uma vez"""
    if not request.transaction_ids:
        raise HTTPException(status_code=400, detail="Nenhuma transação selecionada")
    
    if not request.update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    
    # Filtrar campos válidos
    valid_fields = ['debit_account', 'credit_account', 'status']
    update_data = {k: v for k, v in request.update_data.items() if k in valid_fields and v}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo válido para atualizar")
    
    # Atualizar todas as transações
    result = await db.transactions.update_many(
        {"id": {"$in": request.transaction_ids}},
        {"$set": update_data}
    )
    
    # Se tiver conta de débito e crédito, salvar no histórico de aprendizado
    if update_data.get('debit_account') and update_data.get('credit_account'):
        # Buscar as transações atualizadas para salvar no histórico
        updated_transactions = await db.transactions.find(
            {"id": {"$in": request.transaction_ids}},
            {"_id": 0}
        ).to_list(len(request.transaction_ids))
        
        # Agrupar por statement_id para obter company_id
        statement_ids = set(t['statement_id'] for t in updated_transactions)
        statements = {}
        for st_id in statement_ids:
            st = await db.bank_statements.find_one({"id": st_id}, {"_id": 0})
            if st:
                statements[st_id] = st
        
        # Salvar histórico para cada descrição única
        descriptions_saved = set()
        for trans in updated_transactions:
            statement = statements.get(trans['statement_id'])
            if not statement:
                continue
            
            company_id = statement['company_id']
            description = trans['description']
            
            # Evitar duplicatas
            key = f"{company_id}:{description}"
            if key in descriptions_saved:
                continue
            descriptions_saved.add(key)
            
            # Verificar se já existe
            existing = await db.classification_history.find_one({
                "company_id": company_id,
                "description_pattern": description
            }, {"_id": 0})
            
            if existing:
                await db.classification_history.update_one(
                    {"id": existing['id']},
                    {
                        "$set": {
                            "debit_account": update_data['debit_account'],
                            "credit_account": update_data['credit_account'],
                            "last_used": datetime.now(timezone.utc).isoformat()
                        },
                        "$inc": {"usage_count": 1}
                    }
                )
            else:
                history_entry = ClassificationHistory(
                    company_id=company_id,
                    description_pattern=description,
                    transaction_type=trans.get('transaction_type', 'D'),
                    debit_account=update_data['debit_account'],
                    credit_account=update_data['credit_account']
                )
                history_doc = history_entry.model_dump()
                history_doc['created_at'] = history_doc['created_at'].isoformat()
                history_doc['last_used'] = history_doc['last_used'].isoformat()
                await db.classification_history.insert_one(history_doc)
        
        logger.info(f"Histórico de classificação atualizado para {len(descriptions_saved)} descrições únicas")
    
    return {
        "message": f"{result.modified_count} transações atualizadas com sucesso",
        "updated_count": result.modified_count
    }

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(transaction_id: str, update: TransactionUpdate):
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    
    trans = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    
    # Se a transação foi classificada manualmente, salvar no histórico de aprendizado
    if trans and update.debit_account and update.credit_account:
        # Buscar o statement para obter company_id
        statement = await db.bank_statements.find_one({"id": trans['statement_id']}, {"_id": 0})
        
        if statement:
            company_id = statement['company_id']
            description = trans['description']
            trans_type = trans['transaction_type']
            
            # Verificar se já existe um histórico com descrição similar
            existing_history = await db.classification_history.find_one({
                "company_id": company_id,
                "description_pattern": description
            }, {"_id": 0})
            
            if existing_history:
                # Atualizar registro existente
                await db.classification_history.update_one(
                    {"id": existing_history['id']},
                    {
                        "$set": {
                            "debit_account": update.debit_account,
                            "credit_account": update.credit_account,
                            "last_used": datetime.now(timezone.utc).isoformat()
                        },
                        "$inc": {"usage_count": 1}
                    }
                )
                logger.info(f"Histórico de classificação atualizado para: {description}")
            else:
                # Criar novo registro no histórico
                history_entry = ClassificationHistory(
                    company_id=company_id,
                    description_pattern=description,
                    transaction_type=trans_type,
                    debit_account=update.debit_account,
                    credit_account=update.credit_account
                )
                history_doc = history_entry.model_dump()
                history_doc['created_at'] = history_doc['created_at'].isoformat()
                history_doc['last_used'] = history_doc['last_used'].isoformat()
                await db.classification_history.insert_one(history_doc)
                logger.info(f"Novo histórico de classificação criado para: {description}")
    
    return trans

@api_router.get("/bank-statements/{statement_id}/export")
async def export_statement(statement_id: str):
    # Buscar statement e transações
    statement = await db.bank_statements.find_one({"id": statement_id}, {"_id": 0})
    if not statement:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    
    transactions = await db.transactions.find({"statement_id": statement_id}, {"_id": 0}).to_list(1000)
    
    # Buscar empresa
    company = await db.companies.find_one({"id": statement['company_id']}, {"_id": 0})
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    
    # Cabeçalhos na ordem correta: Descrição, Data, Valor, Conta Débito, Conta Crédito, Histórico
    headers = ['Descrição', 'Data', 'Valor', 'Conta Débito', 'Conta Crédito', 'Histórico']
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar dados
    for trans in transactions:
        valor = trans['amount'] if trans['transaction_type'] == 'C' else -trans['amount']
        ws.append([
            'IMPLANTAÇÃO DE SALDO',  # Descrição fixa
            trans['date'],
            valor,
            trans.get('debit_account', ''),
            trans.get('credit_account', ''),
            trans['description']  # Histórico é a descrição original
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25  # Descrição
    ws.column_dimensions['B'].width = 12  # Data
    ws.column_dimensions['C'].width = 15  # Valor
    ws.column_dimensions['D'].width = 15  # Conta Débito
    ws.column_dimensions['E'].width = 15  # Conta Crédito
    ws.column_dimensions['F'].width = 50  # Histórico
    
    # Salvar arquivo em formato XLSX
    cnpj_clean = clean_cnpj(company['cnpj'])
    filename = f"{cnpj_clean}_{statement['bank_name'].upper()}_{statement['period'].replace('/', '')}_{statement_id[:8]}_LANCAMENTOS.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api_router.delete("/bank-statements/{statement_id}")
async def delete_statement(statement_id: str):
    result = await db.bank_statements.delete_one({"id": statement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Extrato não encontrado")
    await db.transactions.delete_many({"statement_id": statement_id})
    return {"message": "Extrato excluído com sucesso"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()